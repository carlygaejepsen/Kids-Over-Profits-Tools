"""
North Carolina MHLCS Public Records Scraper

Loads the starter NC youth facilities workbook, matches each licensed facility
to the public records directory on the NC DHHS MHLCS site, and OCRs every
inspection/report PDF linked from the facility page before posting the
normalized facility/report payloads to the Kids Over Profits inspections API.
"""

from __future__ import annotations

import argparse
import logging
import os
import re
import time
from collections import defaultdict
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from pdf2image import convert_from_bytes
import pytesseract
from openpyxl import load_workbook

from inspection_api_client import post_facilities_to_api
from scraper_state import load_state, merge_new_ids, save_state, seen_from_state

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

API_URL = os.getenv(
    "INSPECTIONS_API_URL",
    "https://kidsoverprofits.org/wp-content/themes/child/api/inspections-write.php",
)
API_KEY = os.getenv("KOP_DATA_API_KEY", "CHANGE_ME")
STATE_FILE = Path(os.getenv("NC_STATE_FILE", ".nc_state.json"))

WORKBOOK_FALLBACKS = [
    Path.cwd() / "nc_youth_facilities.xlsx",
    Path.cwd() / "nc_youth_facilities.csv",
    Path(__file__).with_name("nc_youth_facilities.xlsx"),
    Path(__file__).with_name("nc_youth_facilities.csv"),
    Path(__file__).resolve().parents[2] / "Kids-Over-Profits" / "nc_youth_facilities.xlsx",
    Path(__file__).resolve().parents[2] / "Kids-Over-Profits" / "nc_youth_facilities.csv",
]

RESULTS_URL = "https://info.ncdhhs.gov/dhsr/mhlcs/sods/results.asp"
GOOGLE_DRIVE_BASE = Path(r"H:\My Drive\FileBird Cloud - kidsoverprofits.org")


def _default_cache_dir(env_name: str, preferred_subdir: str, local_fallback: str) -> Path:
    env_value = os.getenv(env_name)
    if env_value:
        return Path(env_value)

    preferred = GOOGLE_DRIVE_BASE / preferred_subdir
    if GOOGLE_DRIVE_BASE.exists():
        return preferred

    return Path(local_fallback)


PDF_CACHE_DIR = _default_cache_dir("NC_PDF_CACHE", "nc_pdfs", ".nc_pdf_cache")
OCR_CACHE_DIR = _default_cache_dir("NC_OCR_CACHE", "nc_ocr", ".nc_ocr_cache")

TESSERACT_CMD = os.getenv("TESSERACT_CMD")
POPPLER_PATH = os.getenv("POPPLER_PATH")
OCR_DPI = int(os.getenv("NC_OCR_DPI", "175"))
if TESSERACT_CMD:
    pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD

if POPPLER_PATH:
    logger.info("Using POPPLER_PATH=%s", POPPLER_PATH)

UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) KOP-NC-scraper"


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    if not text or text.lower() == "none":
        return ""
    return re.sub(r"\s+", " ", text)


def norm(value: object) -> str:
    text = clean_text(value).lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    for token in (" llc", " inc", " ltd", " company", " corporation", " corp", " llp", " pc"):
        text = text.replace(token, " ")
    return re.sub(r"\s+", " ", text).strip()


def format_date(value: object) -> str:
    text = clean_text(value)
    if not text:
        return ""
    for pattern in ("%d-%b-%y", "%d-%b-%Y", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, pattern).strftime("%m/%d/%Y")
        except ValueError:
            continue
    return text


def find_source_file(explicit_path: Optional[str]) -> Path:
    if explicit_path:
        path = Path(explicit_path).expanduser().resolve()
        if path.exists():
            return path
        raise FileNotFoundError(path)

    for candidate in WORKBOOK_FALLBACKS:
        if candidate.exists():
            return candidate.resolve()

    raise FileNotFoundError("Could not find nc_youth_facilities.xlsx. Pass --input or set NC_SOURCE_FILE.")


def load_workbook_rows(source_file: Path) -> List[Dict[str, object]]:
    if source_file.suffix.lower() == ".csv":
        import csv

        with source_file.open("r", encoding="utf-8-sig", newline="") as handle:
            return [dict(row) for row in csv.DictReader(handle)]

    workbook = load_workbook(source_file, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)

    try:
        headers = [clean_text(value) for value in next(rows)]
    except StopIteration:
        return []

    records: List[Dict[str, object]] = []
    for raw_row in rows:
        record: Dict[str, object] = {}
        for header, value in zip(headers, raw_row):
            if header:
                record[header] = value
        records.append(record)
    return records


def group_workbook_rows(rows: Iterable[Dict[str, object]]) -> Dict[str, List[Dict[str, object]]]:
    grouped: Dict[str, List[Dict[str, object]]] = defaultdict(list)
    for row in rows:
        license_number = clean_text(row.get("License #"))
        if license_number:
            grouped[license_number].append(row)
    return grouped


def pick_representative_row(rows: List[Dict[str, object]]) -> Dict[str, object]:
    def completeness(row: Dict[str, object]) -> int:
        return sum(1 for value in row.values() if clean_text(value))

    return max(rows, key=completeness)


def fetch_directory(session: requests.Session) -> List[Dict[str, str]]:
    response = session.get(RESULTS_URL, timeout=60)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, "html.parser")

    entries: List[Dict[str, str]] = []
    for anchor in soup.select('a[href*="facility.asp?fid="]'):
        row = anchor.find_parent("tr")
        if not row:
            continue
        cells = [cell.get_text(" ", strip=True) for cell in row.find_all("td")]
        if len(cells) < 4:
            continue
        href = urljoin(RESULTS_URL, anchor.get("href", ""))
        fid_match = re.search(r"fid=(\d+)", href)
        if not fid_match:
            continue
        entries.append(
            {
                "fid": fid_match.group(1),
                "name": cells[0],
                "address": cells[1],
                "city": cells[2],
                "zip": cells[3],
                "url": href,
            }
        )
    return entries


def score_match(record: Dict[str, object], entry: Dict[str, str]) -> float:
    candidate_names = [clean_text(record.get("DBA Name")), clean_text(record.get("Name of Licensee Legal Name"))]
    candidate_addresses = [
        clean_text(record.get("Site Address")),
        clean_text(record.get("Facility Address")),
        clean_text(record.get("Site City")),
        clean_text(record.get("Facility City")),
    ]
    candidate_name_blob = norm(" ".join(value for value in candidate_names if value))
    candidate_address_blob = norm(
        " ".join(
            value
            for value in [
                clean_text(record.get("Site Address")),
                clean_text(record.get("Site City")),
                clean_text(record.get("Site Zip")),
                clean_text(record.get("Facility Address")),
                clean_text(record.get("Facility City")),
                clean_text(record.get("Facility Zip")),
            ]
            if value
        )
    )

    entry_name = norm(entry["name"])
    entry_address = norm(f"{entry['address']} {entry['city']} {entry['zip']}")

    best = 0.0
    for cand in (candidate_name_blob, candidate_address_blob, *[norm(value) for value in candidate_names + candidate_addresses]):
        if not cand:
            continue
        if cand == entry_name:
            best = max(best, 100.0)
        elif cand == entry_address:
            best = max(best, 97.0)
        elif cand in entry_name or entry_name in cand:
            best = max(best, 94.0)
        elif cand in entry_address or entry_address in cand:
            best = max(best, 91.0)
        else:
            ratio = max(
                SequenceMatcher(None, cand, entry_name).ratio(),
                SequenceMatcher(None, cand, entry_address).ratio(),
            )
            if ratio >= 0.86:
                best = max(best, 85.0 + (ratio * 15.0))

    county = norm(record.get("County "))
    if county and county in entry_address:
        best += 1.0

    return best


def match_directory_entry(record: Dict[str, object], entries: List[Dict[str, str]]) -> Optional[Dict[str, str]]:
    if not entries:
        return None

    scored = [(score_match(record, entry), entry) for entry in entries]
    scored.sort(key=lambda item: item[0], reverse=True)
    score, entry = scored[0]
    if score < 80.0:
        return None

    return entry


def parse_facility_metadata(soup: BeautifulSoup) -> Dict[str, str]:
    facility_heading = soup.find("h3")
    facility_name = facility_heading.get_text(" ", strip=True) if facility_heading else ""

    program_rows = soup.find_all("table")
    services = ""
    facility_type = ""
    disability_category = ""
    if len(program_rows) > 1:
        data_rows = program_rows[1].find_all("tr")
        if len(data_rows) > 1:
            cells = [cell.get_text(" ", strip=True) for cell in data_rows[1].find_all(["td", "th"])]
            if len(cells) >= 5:
                services = cells[1]
                facility_type = cells[3]
                disability_category = cells[4]

    contact_text = soup.get_text(" ", strip=True)
    contact_match = re.search(r"In Care of:\s*(.*?)\s*Phone:\s*\(?([0-9\-\)\(\s]+)", contact_text)
    contact_name = clean_text(contact_match.group(1)) if contact_match else ""
    phone = clean_text(contact_match.group(2)) if contact_match else ""

    address_match = re.search(
        r"Facility Address\s*(.*?)\s*Mailing Address",
        contact_text,
    )
    facility_address = clean_text(address_match.group(1)) if address_match else ""

    mailing_match = re.search(r"Mailing Address\s*(.*?)\s*Contact Information", contact_text)
    mailing_address = clean_text(mailing_match.group(1)) if mailing_match else ""

    county_match = re.search(r"([A-Za-z ]+ County)", contact_text)
    county = clean_text(county_match.group(1)) if county_match else ""

    return {
        "facility_name": facility_name,
        "services": services,
        "facility_type": facility_type,
        "disability_category": disability_category,
        "contact_name": contact_name,
        "phone": phone,
        "facility_address": facility_address,
        "mailing_address": mailing_address,
        "county": county,
    }


def fetch_pdf_bytes(session: requests.Session, pdf_url: str) -> Optional[bytes]:
    if not pdf_url:
        return None

    PDF_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    cache_name = Path(urlparse(pdf_url).path).name
    cache_path = PDF_CACHE_DIR / cache_name
    if cache_path.exists():
        try:
            data = cache_path.read_bytes()
        except OSError as exc:
            logger.warning("  PDF cache unreadable (%s), re-downloading: %s", exc, cache_name)
            try:
                cache_path.unlink()
            except OSError:
                pass
        else:
            if data.startswith(b"%PDF"):
                logger.info("  PDF cache hit: %s (%d KB)", cache_name, len(data) // 1024)
                return data
            logger.warning("  PDF cache invalid (%d bytes, not a PDF), re-downloading: %s", len(data), cache_name)
            try:
                cache_path.unlink()
            except OSError:
                pass

    logger.info("  PDF download start: %s", pdf_url)
    start = time.monotonic()
    try:
        response = session.get(pdf_url, timeout=120)
        response.raise_for_status()
    except requests.RequestException as exc:
        logger.warning("  PDF download failed for %s after %.1fs: %s", pdf_url, time.monotonic() - start, exc)
        return None

    if not response.content.startswith(b"%PDF"):
        logger.warning("  Non-PDF response for %s", pdf_url)
        return None

    cache_path.write_bytes(response.content)
    logger.info("  PDF downloaded: %s (%d KB, %.1fs)", cache_name, len(response.content) // 1024, time.monotonic() - start)
    return response.content


def ocr_pdf_bytes(pdf_bytes: bytes, cache_key: str) -> str:
    OCR_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    cache_path = OCR_CACHE_DIR / f"{cache_key}.txt"
    if cache_path.exists():
        try:
            text = cache_path.read_text(encoding="utf-8", errors="replace")
        except OSError as exc:
            logger.warning("  OCR cache unreadable (%s), re-running OCR: %s", exc, cache_key)
            try:
                cache_path.unlink()
            except OSError:
                pass
        else:
            logger.info("  OCR cache hit: %s", cache_key)
            return text

    logger.info("  OCR rasterize start: %s (dpi=%d, %d KB)", cache_key, OCR_DPI, len(pdf_bytes) // 1024)
    raster_start = time.monotonic()
    images = convert_from_bytes(pdf_bytes, dpi=OCR_DPI, poppler_path=POPPLER_PATH or None)
    logger.info("  OCR rasterized %d pages in %.1fs: %s", len(images), time.monotonic() - raster_start, cache_key)

    text_parts: List[str] = []
    ocr_start = time.monotonic()
    for page_num, image in enumerate(images, start=1):
        page_start = time.monotonic()
        try:
            page_text = pytesseract.image_to_string(image)
        except Exception as exc:
            logger.warning("  OCR failed for %s page %d: %s", cache_key, page_num, exc)
            break
        page_elapsed = time.monotonic() - page_start
        if page_elapsed > 15.0:
            logger.warning("  OCR slow page: %s page %d took %.1fs", cache_key, page_num, page_elapsed)
        else:
            logger.info("  OCR page %d/%d done (%.1fs, %d chars): %s", page_num, len(images), page_elapsed, len(page_text), cache_key)
        text_parts.append(page_text)

    logger.info("  OCR total %.1fs (%d pages): %s", time.monotonic() - ocr_start, len(text_parts), cache_key)

    text = "\n\n".join(part.strip() for part in text_parts if part).strip()
    if text:
        cache_path.write_text(text, encoding="utf-8")
    return text


def parse_reports(session: requests.Session, facility_url: str, fid: str) -> List[Dict[str, object]]:
    response = session.get(facility_url, timeout=60)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, "html.parser")

    tables = soup.find_all("table")
    if len(tables) < 3:
        return []

    report_rows = tables[2].find_all("tr")
    reports: List[Dict[str, object]] = []
    pdf_rows = [row for row in report_rows[1:] if row.find("a", href=True)]
    logger.info("  fid=%s has %d report rows", fid, len(pdf_rows))

    for pdf_index, row in enumerate(report_rows[1:], start=1):
        cells = [cell.get_text(" ", strip=True) for cell in row.find_all(["td", "th"])]
        if len(cells) < 4:
            continue

        link = row.find("a", href=True)
        if not link:
            continue

        pdf_url = urljoin(facility_url, link.get("href", ""))
        pdf_name = Path(urlparse(pdf_url).path).name or f"{fid}-{cells[2]}"
        report_id = re.sub(r"[^A-Za-z0-9._-]+", "-", pdf_name)
        logger.info("  report %d/%d (fid=%s): %s", pdf_index, len(pdf_rows), fid, report_id)
        report_start = time.monotonic()
        pdf_bytes = fetch_pdf_bytes(session, pdf_url)
        ocr_text = ocr_pdf_bytes(pdf_bytes, report_id) if pdf_bytes else ""
        logger.info("  report %d/%d done in %.1fs (%d OCR chars): %s", pdf_index, len(pdf_rows), time.monotonic() - report_start, len(ocr_text), report_id)

        report_date = cells[2]
        inspection_type = cells[0]
        document_type = cells[1]
        pages = cells[3]

        summary = f"{document_type} - {inspection_type}".strip(" -")

        reports.append(
            {
                "report_id": report_id,
                "report_date": report_date,
                "raw_content": ocr_text,
                "content_length": len(ocr_text),
                "summary": summary,
                "categories": {
                    "inspection_type": inspection_type,
                    "document_type": document_type,
                    "inspection_date": report_date,
                    "pages": pages,
                    "pdf_url": pdf_url,
                    "fid": fid,
                },
            }
        )

    return reports


def build_facility_payload(record: Dict[str, object], entry: Dict[str, str], metadata: Dict[str, str], reports: List[Dict[str, object]]) -> Dict[str, object]:
    program_codes = sorted({clean_text(row.get("Program Code")) for row in [record] if clean_text(row.get("Program Code"))})
    program_code_type = clean_text(record.get("Program Code Type"))
    facility_type = metadata.get("facility_type") or clean_text(record.get("Facility Type"))

    return {
        "facility_info": {
            "facility_name": metadata.get("facility_name") or clean_text(record.get("DBA Name")) or clean_text(record.get("Name of Licensee Legal Name")),
            "program_name": clean_text(record.get("License #")),
            "program_category": facility_type or program_code_type,
            "full_address": metadata.get("facility_address") or _join_address(record.get("Site Address"), record.get("Site City"), record.get("Site State"), record.get("Site Zip")),
            "phone": metadata.get("phone") or clean_text(record.get("Facility Contact Number")),
            "bed_capacity": clean_text(record.get("Beds")) or clean_text(record.get("Total Bed Count")),
            "executive_director": metadata.get("contact_name") or clean_text(record.get("Facility Contact Name")),
            "license_exp_date": format_date(record.get("Expiry Date")),
            "relicense_visit_date": "",
            "action": "Licensed",
        },
        "reports": reports,
        "source": {
            "fid": entry["fid"],
            "public_records_url": entry["url"],
            "county": metadata.get("county") or clean_text(record.get("County ")),
            "services": metadata.get("services"),
            "disability_category": metadata.get("disability_category"),
            "workbook_program_codes": program_codes,
        },
    }


def _join_address(*parts: object) -> str:
    values = [clean_text(part) for part in parts if clean_text(part)]
    return " ".join(values)


def scrape(source_file: Path, limit: Optional[int] = None, full: bool = False) -> Tuple[List[Dict[str, object]], Dict[str, List[str]], List[Dict[str, object]]]:
    session = requests.Session()
    session.headers.update({"User-Agent": UA})

    rows = load_workbook_rows(source_file)
    grouped_rows = group_workbook_rows(rows)
    directory_entries = fetch_directory(session)
    logger.info("Loaded %d workbook rows (%d license groups) and %d directory entries", len(rows), len(grouped_rows), len(directory_entries))

    state = load_state(STATE_FILE)
    seen = {} if full else seen_from_state(state)
    new_ids: Dict[str, List[str]] = {}

    facilities: List[Dict[str, object]] = []
    unmatched: List[Dict[str, object]] = []

    for index, (license_number, group_rows) in enumerate(sorted(grouped_rows.items()), start=1):
        representative = pick_representative_row(group_rows)
        entry = match_directory_entry(representative, directory_entries)
        if not entry:
            unmatched.append(
                {
                    "license_number": license_number,
                    "facility_name": clean_text(representative.get("DBA Name")) or clean_text(representative.get("Name of Licensee Legal Name")),
                    "site_address": clean_text(representative.get("Site Address")),
                    "county": clean_text(representative.get("County ")),
                }
            )
            continue

        facility_url = entry["url"]
        fid = entry["fid"]
        seen_for_fid = seen.get(fid, set())

        logger.info("[%s/%s] %s (fid=%s) - fetching facility page", index, license_number, entry["name"], fid)
        facility_start = time.monotonic()
        try:
            metadata = parse_facility_metadata(BeautifulSoup(session.get(facility_url, timeout=60).text, "html.parser"))
            reports = parse_reports(session, facility_url, fid)
        except requests.RequestException as exc:
            logger.warning("[%s/%s] failed to fetch facility page: %s", index, license_number, exc)
            continue
        logger.info("[%s/%s] %s (fid=%s) - facility processing took %.1fs, %d reports parsed", index, license_number, entry["name"], fid, time.monotonic() - facility_start, len(reports))

        new_reports = [report for report in reports if report["report_id"] and report["report_id"] not in seen_for_fid]
        if not new_reports:
            logger.info("[%s/%s] %s - no new OCR reports", index, license_number, entry["name"])
            continue

        facility_payload = build_facility_payload(representative, entry, metadata, new_reports)
        facilities.append(facility_payload)
        new_ids[fid] = [report["report_id"] for report in new_reports]

        logger.info(
            "[%s/%s] %s - %d OCR reports",
            index,
            license_number,
            entry["name"],
            len(new_reports),
        )

        if limit is not None and len(facilities) >= limit:
            break

    return facilities, new_ids, unmatched


def save_to_api(facilities: List[Dict[str, object]], api_url: str) -> bool:
    result = post_facilities_to_api(
        api_url=api_url,
        api_key=API_KEY,
        state="NC",
        scraped_timestamp=datetime.now().isoformat(),
        facilities=facilities,
        timeout=180,
        info=logger.info,
        error=logger.error,
    )
    return bool(result.get("success"))


def main() -> None:
    parser = argparse.ArgumentParser(description="North Carolina MHLCS public records OCR scraper")
    parser.add_argument("--input", help="Path to nc_youth_facilities.xlsx or a CSV export of it")
    parser.add_argument("--api-url", default=API_URL, help="Override the inspections write endpoint")
    parser.add_argument("--limit", type=int, help="Only post the first N matched facilities")
    parser.add_argument("--no-post", action="store_true", help="Parse, match, and OCR but do not post to the API")
    parser.add_argument("--full", action="store_true", help="Ignore the saved seen-state and reprocess all matched reports")
    args = parser.parse_args()

    source_file = find_source_file(args.input)
    logger.info("Using NC source file: %s", source_file)

    facilities, new_ids, unmatched = scrape(source_file, limit=args.limit, full=args.full)

    logger.info("Matched %d facilities with OCR reports", len(facilities))
    if unmatched:
        logger.info("Unmatched workbook rows: %d", len(unmatched))
        logger.info("First unmatched sample: %s", unmatched[0])

    if args.no_post:
        logger.info("--no-post supplied; skipping API write")
        return

    if not facilities:
        logger.info("No new OCR reports found")
        return

    logger.info("Posting %d facilities to the API", len(facilities))
    if save_to_api(facilities, api_url=args.api_url):
        state = load_state(STATE_FILE)
        merge_new_ids(state, new_ids)
        save_state(STATE_FILE, state)
        logger.info("NC OCR scrape saved successfully")
    else:
        logger.error("API save failed")


if __name__ == "__main__":
    main()